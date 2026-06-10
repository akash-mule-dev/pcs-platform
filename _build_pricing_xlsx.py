"""Generate PCS-Pricing-Model.xlsx — a live, formula-driven pricing & ARR model.
Edit the blue cells on the Assumptions sheets; ARR + Unit Economics recompute.
Two markets: US (USD) and India (INR), each with its own Assumptions / ARR / Unit-Econ tabs.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---- styles -------------------------------------------------------------
TITLE   = Font(bold=True, size=14, color="1F4E78")
HEAD    = Font(bold=True, color="FFFFFF")
SECTION = Font(bold=True, size=11, color="1F4E78")
BOLD    = Font(bold=True)
HEAD_FILL  = PatternFill("solid", fgColor="1F4E78")
INPUT_FILL = PatternFill("solid", fgColor="DDEBF7")   # blue = editable
CALC_FILL  = PatternFill("solid", fgColor="F2F2F2")   # grey = computed
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

USD  = '#,##0'
USD0 = '$#,##0'
INR  = '"₹"#,##0'
PCT  = '0.0%'
NUM  = '0'
MON  = '0.0" mo"'
RAT  = '0.0"x"'
MULT = '0.0"x"'

def cell(ws, ref, value, *, font=None, fill=None, fmt=None, align=None, border=False):
    c = ws[ref]
    c.value = value
    if font: c.font = font
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    if align: c.alignment = Alignment(horizontal=align)
    if border: c.border = BORDER
    return c

# =========================================================================
# Inputs per market (identical CELL LAYOUT, different values)
# =========================================================================
US = dict(
    cur=USD0, core=12000, qual=30000, dt=54000, seat=720,
    smb_acv=22000, mid_acv=150000, smb_onb=7500, mid_onb=20000,
    smb_nrr=1.00, mid_nrr=1.15, smb_churn=0.15, mid_churn=0.08,
    plan_smb=(12, 25, 40), plan_mid=(2, 5, 9), mult=(0.6, 1.0, 1.5),
    smb_cac=8000, mid_cac=45000, onb_gm=0.40,
    host=1200, cad=1500, support=2000, pay=0.025, smb_sites=1, mid_sites=4,
    smb_has_cad=True,   # US blended model charges CAD to all sites (conservative)
)
IN = dict(
    cur=INR, core=144000, qual=360000, dt=600000, seat=10800,
    smb_acv=250000, mid_acv=1750000, smb_onb=50000, mid_onb=350000,
    smb_nrr=0.95, mid_nrr=1.10, smb_churn=0.20, mid_churn=0.10,
    plan_smb=(20, 50, 90), plan_mid=(3, 8, 15), mult=(0.6, 1.0, 1.5),
    smb_cac=125000, mid_cac=800000, onb_gm=0.40,
    host=35000, cad=100000, support=20000, pay=0.02, smb_sites=1, mid_sites=4,
    smb_has_cad=False,  # India: SMB buys non-3D tiers, so no CAD compute COGS
)

# =========================================================================
# Builders
# =========================================================================
def build_assumptions(a, V, currency_label):
    a.sheet_view.showGridLines = False
    cur = V["cur"]
    cell(a, "A1", f"Model Assumptions ({currency_label}) — edit the blue cells", font=TITLE)

    def label(r, t): cell(a, f"A{r}", t)
    def section(r, t): cell(a, f"A{r}", t, font=SECTION)
    def inp(ref, v, fmt=None): cell(a, ref, v, fill=INPUT_FILL, fmt=fmt or cur, border=True, align="right")

    section(3, "PRICING (per site / year)")
    label(4, "Core MES");                inp("B4", V["core"])
    label(5, "Quality & Traceability");  inp("B5", V["qual"])
    label(6, "Digital Twin / AR");       inp("B6", V["dt"])
    label(7, "Office/admin seat / yr");  inp("B7", V["seat"])

    section(9, "SEGMENT ECONOMICS")
    label(10, "SMB blended ACV");           inp("B10", V["smb_acv"])
    label(11, "Mid-market blended ACV");    inp("B11", V["mid_acv"])
    label(12, "SMB onboarding (one-time)"); inp("B12", V["smb_onb"])
    label(13, "Mid onboarding (one-time)"); inp("B13", V["mid_onb"])

    section(15, "RETENTION")
    label(16, "SMB net revenue retention"); inp("B16", V["smb_nrr"], PCT)
    label(17, "Mid NRR");                   inp("B17", V["mid_nrr"], PCT)
    label(18, "SMB logo churn / yr");       inp("B18", V["smb_churn"], PCT)
    label(19, "Mid logo churn / yr");       inp("B19", V["mid_churn"], PCT)

    section(21, "NEW-LOGO PLAN (Base scenario)")
    for col, y in (("B", "Y1"), ("C", "Y2"), ("D", "Y3")):
        cell(a, f"{col}22", y, font=BOLD, align="right")
    label(23, "SMB new logos")
    label(24, "Mid new logos")
    for i, col in enumerate(("B", "C", "D")):
        inp(f"{col}23", V["plan_smb"][i], NUM)
        inp(f"{col}24", V["plan_mid"][i], NUM)

    section(26, "SCENARIO MULTIPLIERS")
    for col, y in (("B", "Low"), ("C", "Base"), ("D", "High")):
        cell(a, f"{col}27", y, font=BOLD, align="right")
    label(28, "× Base plan")
    for i, col in enumerate(("B", "C", "D")):
        inp(f"{col}28", V["mult"][i], MULT)

    section(30, "UNIT ECONOMICS INPUTS")
    label(31, "SMB CAC per logo");        inp("B31", V["smb_cac"])
    label(32, "Mid CAC per logo");        inp("B32", V["mid_cac"])
    label(33, "Onboarding gross margin %"); inp("B33", V["onb_gm"], PCT)

    section(35, "COGS per SITE / year")
    label(36, "Hosting & infra (USD-priced)");       inp("B36", V["host"])
    label(37, "CAD-conversion compute (USD-priced)"); inp("B37", V["cad"])
    label(38, "Support / CS");                        inp("B38", V["support"])
    label(39, "Payment processing (% of ACV)");       inp("B39", V["pay"], PCT)

    section(41, "SITES PER CUSTOMER")
    label(42, "SMB sites"); inp("B42", V["smb_sites"], NUM)
    label(43, "Mid sites"); inp("B43", V["mid_sites"], NUM)

    a.column_dimensions["A"].width = 34
    for col in ("B", "C", "D"):
        a.column_dimensions[col].width = 14


def build_arr(m, assum, cur):
    m.sheet_view.showGridLines = False
    A = f"'{assum}'"
    cell(m, "A1", f"ARR Model — ending ARR by scenario ({assum.split()[0] if ' ' in assum else 'US'})", font=TITLE)

    def hdr_yrs(r):
        cell(m, f"A{r}", "", font=BOLD)
        for col, y in (("B", "Y1"), ("C", "Y2"), ("D", "Y3")):
            cell(m, f"{col}{r}", y, font=HEAD, fill=HEAD_FILL, align="right")

    cell(m, "A3", "ENDING ARR BY SCENARIO", font=SECTION)
    hdr_yrs(4)
    for r, name, src in ((5, "Low", 20), (6, "Base", 32), (7, "High", 44)):
        cell(m, f"A{r}", name, font=BOLD)
        for col in ("B", "C", "D"):
            cell(m, f"{col}{r}", f"={col}{src}", fmt=cur, fill=CALC_FILL, border=True)
    cell(m, "A9", "Services revenue (one-time, Base)", font=BOLD)
    for col in ("B", "C", "D"):
        cell(m, f"{col}9", f"={col}33", fmt=cur, fill=CALC_FILL, border=True)

    for top, mult, title in [(12, "$B$28", "LOW SCENARIO"), (24, "$C$28", "BASE SCENARIO"), (36, "$D$28", "HIGH SCENARIO")]:
        cell(m, f"A{top}", title, font=SECTION)
        hdr_yrs(top + 1)
        sn, mn, sa, ma, se, me, tot, svc = (top + i for i in range(2, 10))
        for nm, r in (("SMB new logos", sn), ("Mid new logos", mn), ("SMB new ARR", sa),
                      ("Mid new ARR", ma), ("SMB ending ARR", se), ("Mid ending ARR", me),
                      ("Total ending ARR", tot), ("Services (one-time)", svc)):
            cell(m, f"A{r}", nm)
        for col in ("B", "C", "D"):
            cell(m, f"{col}{sn}", f"=ROUND({A}!{col}23*{A}!{mult},0)", fmt=NUM, border=True)
            cell(m, f"{col}{mn}", f"=ROUND({A}!{col}24*{A}!{mult},0)", fmt=NUM, border=True)
            cell(m, f"{col}{sa}", f"={col}{sn}*{A}!$B$10", fmt=cur, border=True)
            cell(m, f"{col}{ma}", f"={col}{mn}*{A}!$B$11", fmt=cur, border=True)
        cell(m, f"B{se}", f"=B{sa}", fmt=cur, border=True)
        cell(m, f"C{se}", f"=B{se}*{A}!$B$16+C{sa}", fmt=cur, border=True)
        cell(m, f"D{se}", f"=C{se}*{A}!$B$16+D{sa}", fmt=cur, border=True)
        cell(m, f"B{me}", f"=B{ma}", fmt=cur, border=True)
        cell(m, f"C{me}", f"=B{me}*{A}!$B$17+C{ma}", fmt=cur, border=True)
        cell(m, f"D{me}", f"=C{me}*{A}!$B$17+D{ma}", fmt=cur, border=True)
        for col in ("B", "C", "D"):
            cell(m, f"{col}{tot}", f"={col}{se}+{col}{me}", fmt=cur, font=BOLD, fill=CALC_FILL, border=True)
            cell(m, f"{col}{svc}", f"={col}{sn}*{A}!$B$12+{col}{mn}*{A}!$B$13", fmt=cur, border=True)

    m.column_dimensions["A"].width = 22
    for col in ("B", "C", "D"):
        m.column_dimensions[col].width = 15


def build_unit_econ(u, assum, cur, smb_has_cad):
    u.sheet_view.showGridLines = False
    A = f"'{assum}'"
    cell(u, "A1", f"Unit Economics ({assum.split()[0] if ' ' in assum else 'US'})", font=TITLE)
    cell(u, "E3", "weights", font=BOLD); cell(u, "F3", "n", font=BOLD)
    cell(u, "E4", "SMB logos (Base)"); cell(u, "F4", f"=SUM({A}!B23:D23)", fmt=NUM)
    cell(u, "E5", "Mid logos (Base)");  cell(u, "F5", f"=SUM({A}!B24:D24)", fmt=NUM)

    cell(u, "A3", "Metric", font=HEAD, fill=HEAD_FILL)
    for col, h in (("B", "SMB"), ("C", "Mid"), ("D", "Blended")):
        cell(u, f"{col}3", h, font=HEAD, fill=HEAD_FILL, align="right")

    smb_cad = "=B5*" + A + "!$B$37" if smb_has_cad else "=0"
    rows = [
        (4,  "ACV / yr",                    f"={A}!B10", f"={A}!B11", cur, "wavg"),
        (5,  "Sites per customer",          f"={A}!B42", f"={A}!B43", NUM, None),
        (6,  "COGS — hosting & infra",      f"=B5*{A}!$B$36", f"=C5*{A}!$B$36", cur, None),
        (7,  "COGS — CAD-conversion compute", smb_cad,       f"=C5*{A}!$B$37", cur, None),
        (8,  "COGS — support / CS",         f"=B5*{A}!$B$38", f"=C5*{A}!$B$38", cur, None),
        (9,  "COGS — payments",             f"=B4*{A}!$B$39", f"=C4*{A}!$B$39", cur, None),
        (10, "Total COGS",                  "=SUM(B6:B9)",   "=SUM(C6:C9)",   cur, "wavg"),
        (11, "Gross profit / yr",           "=B4-B10",       "=C4-C10",       cur, "wavg"),
        (12, "Gross margin %",              "=B11/B4",       "=C11/C4",       PCT, "gm"),
        (14, "CAC per logo",                f"={A}!B31",     f"={A}!B32",     cur, "wavg"),
        (15, "Onboarding fee",              f"={A}!B12",     f"={A}!B13",     cur, None),
        (16, "Onboarding gross profit",     f"=B15*{A}!$B$33", f"=C15*{A}!$B$33", cur, None),
        (17, "Net CAC (CAC − onb. GP)",     "=B14-B16",      "=C14-C16",      cur, "wavg"),
        (19, "Monthly gross profit",        "=B11/12",       "=C11/12",       cur, None),
        (20, "CAC payback (gross)",         "=B14/B19",      "=C14/C19",      MON, "pb"),
        (21, "Net CAC payback",             "=B17/B19",      "=C17/C19",      MON, "pbn"),
        (23, "Logo churn / yr",             f"={A}!B18",     f"={A}!B19",     PCT, None),
        (24, "Customer lifetime (yrs)",     "=1/B23",        "=1/C23",        "0.0", None),
        (25, "LTV (gross margin)",          "=B11*B24",      "=C11*C24",      cur, "wavg"),
        (26, "LTV:CAC (gross)",             "=B25/B14",      "=C25/C14",      RAT, "ltvcac"),
        (27, "LTV:CAC (net CAC)",           "=B25/B17",      "=C25/C17",      RAT, "ltvcacn"),
    ]
    emph = ("Gross margin %", "LTV:CAC (net CAC)", "Net CAC payback")
    for r, name, bf, cf, fmt, blend in rows:
        cell(u, f"A{r}", name, font=(BOLD if name in emph else None))
        cell(u, f"B{r}", bf, fmt=fmt, border=True)
        cell(u, f"C{r}", cf, fmt=fmt, border=True)
        d = {
            "wavg": f"=(B{r}*$F$4+C{r}*$F$5)/($F$4+$F$5)",
            "gm": "=D11/D4", "pb": "=D14/(D11/12)", "pbn": "=D17/(D11/12)",
            "ltvcac": "=D25/D14", "ltvcacn": "=D25/D17",
        }.get(blend)
        if d:
            cell(u, f"D{r}", d, fmt=fmt, fill=CALC_FILL, border=True)

    if not smb_has_cad:
        cell(u, "A29", "Note: SMB excludes CAD compute (SMB = non-3D tiers). Mid = Digital Twin tier (CAD applies).", font=Font(italic=True, color="7F7F7F"))
    u.column_dimensions["A"].width = 30
    for col in ("B", "C", "D"):
        u.column_dimensions[col].width = 14
    u.column_dimensions["E"].width = 18
    u.column_dimensions["F"].width = 8


# =========================================================================
# README
# =========================================================================
wb = Workbook()
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
cell(ws, "A1", "PCS Platform — Pricing & ARR Model", font=TITLE)
notes = [
    "", "TABS",
    "  • US Assumptions / US ARR Model / US Unit Economics  — North-American pricing in USD",
    "  • India Assumptions / India ARR Model / India Unit Economics  — India pricing in INR (ex-GST)",
    "", "HOW TO USE",
    "  • Edit only the BLUE cells on an Assumptions tab. Everything else is a formula.",
    "  • ARR rolls each segment's bookings forward by NRR and sums Low/Base/High scenarios.",
    "  • Unit Economics derives gross margin, CAC payback and LTV:CAC; CAD-conversion compute is broken out as COGS.",
    "", "MODEL LOGIC",
    "  • Metric = price per production SITE; floor/operator access unlimited, only office/admin seats counted.",
    "  • Ending ARR(t) = Ending ARR(t-1) × NRR + new-logo ARR(t), per segment then summed.",
    "  • LTV = annual gross profit ÷ logo churn (no expansion credit — conservative).",
    "  • Net CAC = CAC − onboarding gross profit (onboarding fees offset acquisition cost).",
    "", "INDIA NOTES",
    "  • Prices are ex-GST (add 18% GST; B2B customers reclaim it). Customers may deduct ~2% TDS on services.",
    "  • Hosting & CAD-conversion compute are USD-priced (cloud is global) so they DON'T localise down —",
    "    this compresses margin on the Digital Twin tier. Support/CS does localise down.",
    "  • India SMB excludes CAD compute COGS (SMB buys non-3D tiers); Mid = Digital Twin tier, so CAD applies.",
    "", "CAVEATS",
    "  • Figures are planning anchors, NOT committed list prices. Validate willingness-to-pay first.",
    "  • LTV:CAC ratios are gated by the churn assumptions — stress-test churn before trusting them.",
    "  • See PRICING.md for the narrative one-pager and the per-tier feature matrix this model backs.",
]
for i, t in enumerate(notes, start=2):
    f = SECTION if t in ("TABS", "HOW TO USE", "MODEL LOGIC", "INDIA NOTES", "CAVEATS") else None
    cell(ws, f"A{i}", t, font=f)
ws.column_dimensions["A"].width = 112

# US tabs
build_assumptions(wb.create_sheet("US Assumptions"), US, "USD")
build_arr(wb.create_sheet("US ARR Model"), "US Assumptions", US["cur"])
build_unit_econ(wb.create_sheet("US Unit Economics"), "US Assumptions", US["cur"], US["smb_has_cad"])
# India tabs
build_assumptions(wb.create_sheet("India Assumptions"), IN, "INR, ex-GST")
build_arr(wb.create_sheet("India ARR Model"), "India Assumptions", IN["cur"])
build_unit_econ(wb.create_sheet("India Unit Economics"), "India Assumptions", IN["cur"], IN["smb_has_cad"])

wb.save("PCS-Pricing-Model.xlsx")
print("wrote PCS-Pricing-Model.xlsx with US + India tabs")
